# modules/output_writer.py
# Module 4：整理 DataFrame，輸出 CSV 與 Excel

import os
import re
import pandas as pd
from datetime import datetime, date

OUTPUT_DIR = "output"
OUTPUT_COLS = [
    "patent_id", "title", "year", "status",
    "expiry_date", "expiry_source",
    "is_target_drug", "delivery_routes", "indications",
    "fto_risk", "gap_opportunity", "reasoning",
]
RISK_ORDER = {"High": 0, "Medium": 1, "Low": 2}

def clean_excel_string(val):
    """移除 Excel 不支援的控制字元 (如 \x00-\x08, \x0b, \x0c, \x0e-\x1f)"""
    if isinstance(val, str):
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val)
    return val

def save_results(results: list[dict], prefix: str = "gap_analysis") -> str:
    """
    接收分析結果 list，整理成 DataFrame 並輸出 CSV + Excel。
    回傳輸出檔案路徑（CSV）。
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_path = os.path.join(OUTPUT_DIR, f"{prefix}_{timestamp}")

    df = pd.DataFrame(results)

    # 關鍵：對所有字串欄位進行清洗
    df = df.map(clean_excel_string)

    # 確保所有欄位存在
    for col in OUTPUT_COLS:
        if col not in df.columns:
            df[col] = ""

    # list 欄位轉成逗號分隔字串（方便 Excel 閱讀）
    for col in ["delivery_routes", "indications"]:
        df[col] = df[col].apply(
            lambda x: ", ".join(x) if isinstance(x, list) else str(x)
        )

    # 排序：High → Medium → Low，同風險內依年份新到舊
    df["_risk_sort"] = df["fto_risk"].map(RISK_ORDER).fillna(3)
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df.sort_values(["_risk_sort", "year"], ascending=[True, False])
    df = df.drop(columns=["_risk_sort"])

    # ── CSV ──────────────────────────────────────────────────────────────────
    csv_path = f"{base_path}.csv"
    df[OUTPUT_COLS].to_csv(csv_path, index=False, encoding="utf-8-sig")

    # ── Excel（加顏色標示風險等級）────────────────────────────────────────────
    xlsx_path = f"{base_path}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df[OUTPUT_COLS].to_excel(writer, index=False, sheet_name="Gap Analysis")
        wb = writer.book
        ws = writer.sheets["Gap Analysis"]

        from openpyxl.styles import PatternFill, Font
        fills = {
            "High":   PatternFill("solid", fgColor="FFCCCC"),  # 淡紅
            "Medium": PatternFill("solid", fgColor="FFF2CC"),  # 淡黃
            "Low":    PatternFill("solid", fgColor="E2EFDA"),  # 淡綠
        }
        risk_col_idx = OUTPUT_COLS.index("fto_risk") + 1  # openpyxl 從 1 開始

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            risk_cell = row[risk_col_idx - 1]
            fill = fills.get(risk_cell.value)
            if fill:
                for cell in row:
                    cell.fill = fill

        # ── Expiry date conditional formatting ───────────────────────
        # expired=灰, <1yr=黃, active=綠, no data=不上色
        expiry_col_idx = OUTPUT_COLS.index("expiry_date") + 1
        expiry_fills = {
            "expired":    PatternFill("solid", fgColor="D9D9D9"),  # 灰
            "expiring":   PatternFill("solid", fgColor="FFF2CC"),  # 黃（<1yr）
            "active":     PatternFill("solid", fgColor="E2EFDA"),  # 綠
        }
        expiry_font_grey = Font(color="808080")  # 灰色字（expired rows）
        today = date.today()

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            expiry_cell = row[expiry_col_idx - 1]
            val = expiry_cell.value
            if not val or not isinstance(val, str):
                continue
            try:
                expiry_d = date.fromisoformat(val)
            except (ValueError, TypeError):
                continue

            if expiry_d <= today:
                # Expired
                expiry_cell.fill = expiry_fills["expired"]
                expiry_cell.font = expiry_font_grey
            elif (expiry_d - today).days <= 365:
                # Expiring within 1 year
                expiry_cell.fill = expiry_fills["expiring"]
            else:
                # Active
                expiry_cell.fill = expiry_fills["active"]

        # 凍結首行
        ws.freeze_panes = "A2"

        # 自動調整欄寬（簡易版）
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    print(f"  CSV  → {csv_path}")
    print(f"  Excel → {xlsx_path}")
    return csv_path


def print_summary(results: list[dict]) -> None:
    """在 terminal 印出簡易統計摘要。"""
    df = pd.DataFrame(results)
    total = len(df)
    counts = df["fto_risk"].value_counts()

    print("\n" + "=" * 40)
    print(f"  分析完成：共 {total} 筆專利")
    print(f"  🔴 High   : {counts.get('High',   0)}")
    print(f"  🟡 Medium : {counts.get('Medium', 0)}")
    print(f"  🟢 Low    : {counts.get('Low',    0)}")
    print("=" * 40)

    high_risk = df[df["fto_risk"] == "High"]
    if not high_risk.empty:
        print("\n  ⚠️  High risk 專利（需人工精讀）：")
        for _, row in high_risk.iterrows():
            print(f"  - {row['patent_id']}  {row['title'][:60]}")
    print()
